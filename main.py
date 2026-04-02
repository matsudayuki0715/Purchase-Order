#!/usr/bin/env python3
"""
発注書作成ツール - KiQ Robotics
メーカからの見積書PDFを読み込み、発注書（Excel/PDF）を自動生成する
"""

import os
import json
import shutil
import datetime
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import openpyxl

# ============================================================
# 環境変数ロード（.env ファイルから）
# ============================================================

def _load_env():
    env_path = Path(__file__).parent / '.env'
    if env_path.exists():
        with open(env_path, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, _, val = line.partition('=')
                    os.environ.setdefault(key.strip(), val.strip())

_load_env()

TEMPLATE_PATH = Path(__file__).parent / 'template.xlsx'
OUTPUT_DIR    = Path(__file__).parent / 'output'


# ============================================================
# PDF抽出（pdfplumber + 正規表現 / EasyOCR）
# ============================================================

def _parse_with_regex(text: str) -> dict | None:
    """参照プロジェクト(claw_gui.py)と同じ手法：pdfplumberテキスト + 正規表現。
    十分な情報が取れなければ None を返す。"""
    import re

    # maker_name: KiQ・松田以外の会社名を探す
    # 行全体に除外キーワードが含まれる行はスキップし、短い会社名行を優先する
    _EXCLUDE = ('KiQ', 'Robo', 'KiQRobotics', '松田', 'Robotics')
    maker = ''
    for line in text.splitlines():
        line = line.strip()
        if any(ng in line for ng in _EXCLUDE):   # 行全体でチェック
            continue
        for pat in [
            r'(SST[\w\s]*?(?:株式会社|有限会社))',
            r'((?:株式会社|有限会社)\s*\S+)',
            r'(\S+\s*(?:株式会社|有限会社))',
        ]:
            m = re.search(pat, line)
            if m:
                candidate = re.sub(r'\s*(御中|様)\s*$', '', m.group(1).strip())
                if len(candidate) >= 5:          # 短すぎる誤検出を除外
                    maker = candidate
                    break
        if maker:
            break

    # quote_no
    quote_no = ''
    for pat in [r'見積書?番号\s+([\w-]+)', r'見積\s*No\.?\s*([\d-]+)']:
        m = re.search(pat, text)
        if m:
            quote_no = m.group(1).strip()
            break

    # subject
    subject = ''
    for pat in [r'件名[：:]\s*(.+)', r'件\s+名\s+(.+)', r'【(.+?見積り?)】']:
        m = re.search(pat, text)
        if m:
            subject = m.group(1).strip()
            break

    # delivery
    delivery = ''
    for pat in [r'納期[：:]\s*(受注後[^\n]+)', r'製作L/T[：:]\s*([^\n]+)']:
        m = re.search(pat, text)
        if m:
            delivery = m.group(1).strip()
            break

    # items ① クリエイティング形式: ・名称：xxx N 個 単価 金額
    item_pat = re.compile(
        r'・(?:名称|図番)[：:](.+?)\s+(\d+)\s+(?:個|式)\s+([\d,]+)\s+([\d,]+)'
    )
    items = [
        {'description': m.group(1).strip(),
         'quantity':    int(m.group(2)),
         'unit_price':  int(m.group(3).replace(',', ''))}
        for m in item_pat.finditer(text)
    ]

    # items ② クリモト形式: 品名 N 式 単価 金額（行頭に番号がない）
    if not items:
        skip = {'合 計', '合計', '摘 要', '摘要', 'rix-std', '承 認', '確 認', '作 成'}
        pat2 = re.compile(r'^(.{3,40}?)\s+(\d+)\s+式\s+([\d,]+)\s+([\d,]+)', re.MULTILINE)
        for m in pat2.finditer(text):
            name = m.group(1).strip()
            if any(s in name for s in skip) or len(name) < 3:
                continue
            items.append({
                'description': name,
                'quantity':    int(m.group(2)),
                'unit_price':  int(m.group(3).replace(',', '')),
            })

    # notes: 【見積り条件】以降
    notes = ''
    m = re.search(r'【見積り?条件】(.+?)(?=【|$)', text, re.DOTALL)
    if m:
        notes = m.group(1).strip()
    elif '■御見積に関する' in text:
        idx = text.index('■御見積に関する')
        notes = text[idx:].strip()

    # 品目が1件以上あれば成功とみなす
    if items:
        return {
            'maker_name': maker,
            'quote_no':   quote_no,
            'subject':    subject,
            'delivery':   delivery,
            'items':      items,
            'notes':      notes,
        }
    return None


def _ocr_with_easyocr(pdf_path: str) -> str:
    """EasyOCR（無料・ローカル）で画像PDFをテキスト化する。初回のみモデルDL（約500MB）。"""
    import fitz, easyocr, tempfile, os as _os
    doc = fitz.open(pdf_path)
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(2, 2))
    tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    pix.save(tmp.name)
    doc.close()
    try:
        reader  = easyocr.Reader(['ja', 'en'], gpu=False, verbose=False)
        results = reader.readtext(tmp.name, detail=0, paragraph=True)
        return '\n'.join(results)
    finally:
        _os.unlink(tmp.name)


def extract_from_pdf(pdf_path: str) -> dict:
    """
    ① pdfplumber/PyMuPDF でテキスト抽出（参照 claw_gui.py と同じ手法）
    ② 正規表現でパース（オフライン・即時）
    ③ 画像PDF → EasyOCR（無料・ローカル）→ 正規表現
    いずれも失敗した場合は RuntimeError を raise → GUIで手動入力を促す
    """
    # ── Step 1: pdfplumber でテキスト抽出 ──
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            pdf_text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
    except Exception:
        pdf_text = ''

    # pdfplumber で取れなければ PyMuPDF でリトライ
    if len(pdf_text.strip()) < 30:
        import fitz
        doc = fitz.open(pdf_path)
        pdf_text = '\n'.join(page.get_text() for page in doc)
        doc.close()

    # ── Step 2: 正規表現でパース ──
    if len(pdf_text.strip()) >= 30:
        result = _parse_with_regex(pdf_text)
        if result:
            return result
        raise RuntimeError(
            "テキストの読み取りはできましたが、自動解析に失敗しました。\n"
            "右側のフォームに手動で入力してください。"
        )

    # ── Step 3: 画像PDF → EasyOCR → 正規表現 ──
    try:
        ocr_text = _ocr_with_easyocr(pdf_path)
    except Exception as e:
        raise RuntimeError(f"EasyOCRが利用できません: {e}\n手動で入力してください。")

    if len(ocr_text.strip()) < 30:
        raise RuntimeError("OCRでテキストを読み取れませんでした。手動で入力してください。")

    result = _parse_with_regex(ocr_text)
    if result:
        return result

    raise RuntimeError(
        "OCRは成功しましたが、自動解析に失敗しました。\n"
        "右側のフォームに手動で入力してください。"
    )


# ============================================================
# Excel生成
# ============================================================

def create_excel(data: dict, output_path: Path) -> None:
    if not TEMPLATE_PATH.exists():
        raise RuntimeError(f"テンプレートファイルが見つかりません: {TEMPLATE_PATH}")

    shutil.copy(TEMPLATE_PATH, output_path)

    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(output_path)

    ws0 = wb.worksheets[0]  # 設定シート

    order_date = data.get('order_date', datetime.date.today())
    if isinstance(order_date, str):
        try:
            order_date = datetime.date.fromisoformat(order_date)
        except ValueError:
            order_date = datetime.date.today()

    ws0['B7']  = order_date.year
    ws0['B8']  = order_date.month
    ws0['B9']  = order_date.day
    ws0['B10'] = int(data.get('serial_no', 1))
    ws0['B2']  = data.get('maker_name', '')
    ws0['B3']  = ''
    ws0['B4']  = '御中'
    ws0['B13'] = data.get('subject', '')
    ws0['B14'] = data.get('delivery', '')
    ws0['B15'] = data.get('delivery_location', 'あるあるcity2号館小倉B1F')
    ws0['B16'] = data.get('payment_terms', '月末締翌月末支払')
    ws0['B17'] = data.get('quote_no', '')
    ws0['J2']  = data.get('notes', '')

    for row in range(2, 27):
        ws0.cell(row=row, column=5).value = None
        ws0.cell(row=row, column=6).value = None
        ws0.cell(row=row, column=7).value = None

    for i, item in enumerate(data.get('items', [])[:25]):
        r = i + 2
        ws0.cell(row=r, column=5).value = item.get('description', '')
        qty   = item.get('quantity', '')
        price = item.get('unit_price', '')
        ws0.cell(row=r, column=6).value = int(qty)   if qty   else None
        ws0.cell(row=r, column=7).value = int(price) if price else None

    wb.save(output_path)
    wb.close()


# ============================================================
# PDF生成（win32com）
# ============================================================

def excel_to_pdf(excel_path: Path, pdf_path: Path) -> None:
    try:
        import win32com.client
    except ImportError:
        raise RuntimeError("pywin32 が利用できません。`py -m pip install pywin32` を実行してください。")

    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(excel_path.resolve()))
        wb.Worksheets(2).ExportAsFixedFormat(
            Type=0, Filename=str(pdf_path.resolve()),
            Quality=0, IncludeDocProperties=True, IgnorePrintAreas=False,
        )
    finally:
        try: wb.Close(False)
        except Exception: pass
        try: excel.Quit()
        except Exception: pass


# ============================================================
# 出力パス生成
# ============================================================

def make_output_path(data: dict, ext: str) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    order_date = data.get('order_date', datetime.date.today())
    date_str   = order_date.strftime('%Y%m%d') if hasattr(order_date, 'strftime') else str(order_date).replace('-', '')
    serial_str = str(data.get('serial_no', 1)).zfill(2)
    maker      = data.get('maker_name', 'メーカ')
    subject    = data.get('subject', '').replace('/', '').replace('\\', '')[:20]
    filename   = f"{date_str}_{serial_str}_発注書_{maker}様_{subject}{ext}"
    for ch in r'<>:"/\|?*':
        filename = filename.replace(ch, '_')
    return OUTPUT_DIR / filename


# ============================================================
# PDFプレビューウィジェット
# ============================================================

class PdfPreview(ttk.Frame):
    """左パネル：PDFページをレンダリングして表示する"""

    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self._doc       = None
        self._page_num  = 0
        self._photo     = None   # GC防止用の参照保持
        self._render_job = None  # after() のキャンセル用
        self._build()

    def _build(self):
        # ナビゲーションバー（上部）
        nav = ttk.Frame(self)
        nav.pack(fill='x', pady=(0, 2))

        self._prev_btn = ttk.Button(nav, text='◀', width=3, command=self._prev_page)
        self._prev_btn.pack(side='left', padx=2)

        self._page_label = ttk.Label(nav, text='PDF未選択', anchor='center', width=14)
        self._page_label.pack(side='left', expand=True)

        self._next_btn = ttk.Button(nav, text='▶', width=3, command=self._next_page)
        self._next_btn.pack(side='left', padx=2)

        self._prev_btn.configure(state='disabled')
        self._next_btn.configure(state='disabled')

        # キャンバス + スクロールバー
        canvas_frame = ttk.Frame(self)
        canvas_frame.pack(fill='both', expand=True)

        self._canvas = tk.Canvas(canvas_frame, bg='#6b6b6b', cursor='crosshair', highlightthickness=0)
        v_sb = ttk.Scrollbar(canvas_frame, orient='vertical',   command=self._canvas.yview)
        h_sb = ttk.Scrollbar(canvas_frame, orient='horizontal', command=self._canvas.xview)
        self._canvas.configure(yscrollcommand=v_sb.set, xscrollcommand=h_sb.set)

        v_sb.pack(side='right',  fill='y')
        h_sb.pack(side='bottom', fill='x')
        self._canvas.pack(fill='both', expand=True)

        # マウスホイールスクロール
        self._canvas.bind('<MouseWheel>', lambda e: self._canvas.yview_scroll(int(-1*(e.delta/120)), 'units'))

        # パネル幅が変わったらリレンダリング（デバウンス付き）
        self._canvas.bind('<Configure>', self._on_resize)

    # ----------------------------------------------------------------
    # 公開API
    # ----------------------------------------------------------------

    def load(self, pdf_path: str):
        """PDFを読み込んで最初のページを表示する"""
        try:
            import fitz
            self._doc      = fitz.open(pdf_path)
            self._page_num = 0
            self._render()
        except Exception as e:
            self._page_label.configure(text=f'読込エラー')
            self._canvas.delete('all')

    def clear(self):
        self._doc = None
        self._photo = None
        self._canvas.delete('all')
        self._page_label.configure(text='PDF未選択')
        self._prev_btn.configure(state='disabled')
        self._next_btn.configure(state='disabled')

    # ----------------------------------------------------------------
    # 内部処理
    # ----------------------------------------------------------------

    def _render(self):
        if not self._doc:
            return
        try:
            from PIL import Image, ImageTk

            page       = self._doc[self._page_num]
            total      = len(self._doc)
            canvas_w   = self._canvas.winfo_width()
            if canvas_w < 50:
                canvas_w = 480
            # A4ページを横幅にフィット（最大2倍ズーム）
            zoom = min(canvas_w / page.rect.width, 2.0)
            mat  = page.parent.matrix if hasattr(page.parent, 'matrix') else None
            pix  = page.get_pixmap(matrix=__import__('fitz').Matrix(zoom, zoom))
            img  = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)

            self._photo = ImageTk.PhotoImage(img)
            self._canvas.delete('all')
            self._canvas.create_image(0, 0, anchor='nw', image=self._photo)
            self._canvas.configure(scrollregion=(0, 0, pix.width, pix.height))
            self._canvas.yview_moveto(0)

            self._page_label.configure(text=f'{self._page_num + 1} / {total} ページ')
            self._prev_btn.configure(state='normal' if self._page_num > 0         else 'disabled')
            self._next_btn.configure(state='normal' if self._page_num < total - 1 else 'disabled')
        except Exception:
            pass

    def _on_resize(self, _event=None):
        # リサイズが落ち着いてから再描画（150ms デバウンス）
        if self._render_job:
            self.after_cancel(self._render_job)
        self._render_job = self.after(150, self._render)

    def _prev_page(self):
        if self._doc and self._page_num > 0:
            self._page_num -= 1
            self._render()

    def _next_page(self):
        if self._doc and self._page_num < len(self._doc) - 1:
            self._page_num += 1
            self._render()


# ============================================================
# 明細行
# ============================================================

class ItemRow:
    def __init__(self, frame, row_idx, on_change, on_delete, description='', quantity='', unit_price=''):
        self.frame     = frame
        self.row_idx   = row_idx
        self.desc_var  = tk.StringVar(value=str(description))
        self.qty_var   = tk.StringVar(value=str(quantity)   if quantity   else '')
        self.price_var = tk.StringVar(value=str(unit_price) if unit_price else '')
        self.amount_var = tk.StringVar(value='')

        self.no_label     = ttk.Label(frame, width=4, anchor='center')
        self.desc_entry   = ttk.Entry(frame, textvariable=self.desc_var,  width=38)
        self.qty_entry    = ttk.Entry(frame, textvariable=self.qty_var,   width=7,  justify='right')
        self.price_entry  = ttk.Entry(frame, textvariable=self.price_var, width=11, justify='right')
        self.amount_label = ttk.Label(frame, textvariable=self.amount_var, width=11, anchor='e')
        self.del_btn      = ttk.Button(frame, text='✕', width=3, command=on_delete)

        self._on_change = on_change
        self.qty_var.trace_add('write',   self._update)
        self.price_var.trace_add('write', self._update)
        self._do_grid()

    def _do_grid(self):
        r = self.row_idx
        self.no_label.grid    (row=r, column=0, padx=2, pady=1)
        self.desc_entry.grid  (row=r, column=1, padx=2, pady=1, sticky='ew')
        self.qty_entry.grid   (row=r, column=2, padx=2, pady=1)
        self.price_entry.grid (row=r, column=3, padx=2, pady=1)
        self.amount_label.grid(row=r, column=4, padx=2, pady=1)
        self.del_btn.grid     (row=r, column=5, padx=2, pady=1)

    def _update(self, *_):
        try:
            self.amount_var.set(f"{int(self.qty_var.get().replace(',','')) * int(self.price_var.get().replace(',','')):,}")
        except (ValueError, TypeError):
            self.amount_var.set('')
        self._on_change()

    def get_amount(self) -> int:
        try:
            return int(self.qty_var.get().replace(',','')) * int(self.price_var.get().replace(',',''))
        except (ValueError, TypeError):
            return 0

    def to_dict(self) -> dict:
        try: qty   = int(self.qty_var.get().replace(',',''))
        except ValueError: qty = 0
        try: price = int(self.price_var.get().replace(',',''))
        except ValueError: price = 0
        return {'description': self.desc_var.get().strip(), 'quantity': qty, 'unit_price': price}

    def is_empty(self) -> bool:
        # 摘要だけでも書いてあれば転写する（数量・単価なしの説明行を許容）
        return not self.desc_var.get().strip()


# ============================================================
# メインアプリ
# ============================================================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('発注書作成ツール  |  KiQ Robotics')
        self.geometry('1380x820')
        self.resizable(True, True)
        self.configure(bg='#f0f0f0')

        self._item_rows: list[ItemRow] = []
        self._next_item_row = 2

        self._build_ui()
        self._populate_defaults()

    # ----------------------------------------------------------------
    # UI構築
    # ----------------------------------------------------------------

    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use('clam')
        style.configure('TFrame',          background='#f0f0f0')
        style.configure('TLabelframe',     background='#f0f0f0')
        style.configure('TLabelframe.Label', background='#f0f0f0', font=('Yu Gothic UI', 9, 'bold'))
        style.configure('TLabel',          background='#f0f0f0')
        style.configure('TButton',         padding=(6, 4))
        style.configure('Action.TButton',  font=('Yu Gothic UI', 10, 'bold'))

        # ヘッダーバー
        header = tk.Frame(self, bg='#1a3a6b', height=46)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text='  発注書作成ツール',
                 font=('Yu Gothic UI', 14, 'bold'), fg='white', bg='#1a3a6b', anchor='w'
                 ).pack(side='left', padx=12, pady=8)

        # 全体コンテナ
        outer = ttk.Frame(self, padding=(8, 6))
        outer.pack(fill='both', expand=True)

        # ① PDF選択バー（全幅）
        self._build_pdf_bar(outer)

        # ② 左右分割エリア（PanedWindow）
        paned = tk.PanedWindow(outer, orient='horizontal', sashwidth=6,
                               bg='#aaa', sashrelief='flat', bd=0)
        paned.pack(fill='both', expand=True, pady=(4, 4))

        # 左パネル：PDFプレビュー
        left_frame = ttk.Frame(paned, padding=(0, 0, 4, 0))
        self._preview = PdfPreview(left_frame)
        self._preview.pack(fill='both', expand=True)
        paned.add(left_frame, minsize=300, width=560)

        # 右パネル：フォーム（スクロール対応）
        right_frame = ttk.Frame(paned, padding=(4, 0, 0, 0))
        paned.add(right_frame, minsize=350)
        self._build_right_panel(right_frame)

        # ③ ボタンバー（全幅）
        self._build_btn_bar(outer)

    def _build_pdf_bar(self, parent):
        bar = ttk.LabelFrame(parent, text='① メーカ見積書PDF', padding=(8, 5))
        bar.pack(fill='x', pady=(0, 0))

        self.pdf_var = tk.StringVar()
        self.pdf_var.trace_add('write', self._on_pdf_path_changed)

        ttk.Entry(bar, textvariable=self.pdf_var, width=72).pack(side='left', padx=(0, 6))
        ttk.Button(bar, text='参照...', command=self._browse_pdf).pack(side='left', padx=(0, 6))
        ttk.Button(bar, text='  AI抽出  ', command=self._extract_async,
                   style='Action.TButton').pack(side='left')

        self.status_var = tk.StringVar(value='PDFを選択するとプレビューが表示されます。「AI抽出」で自動入力できます。')
        self.status_label = ttk.Label(bar, textvariable=self.status_var,
                                       foreground='#555', font=('Yu Gothic UI', 8))
        self.status_label.pack(side='left', padx=10)

    def _build_right_panel(self, parent):
        """右パネル：スクロール可能なフォーム"""
        canvas = tk.Canvas(parent, bg='#f0f0f0', highlightthickness=0)
        v_sb   = ttk.Scrollbar(parent, orient='vertical', command=canvas.yview)
        self.scroll_frame = ttk.Frame(canvas)

        self.scroll_frame.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )
        win_id = canvas.create_window((0, 0), window=self.scroll_frame, anchor='nw')
        canvas.configure(yscrollcommand=v_sb.set)
        canvas.bind('<Configure>', lambda e: canvas.itemconfig(win_id, width=e.width))
        canvas.bind_all('<MouseWheel>',
                        lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), 'units'))

        canvas.pack(side='left', fill='both', expand=True)
        v_sb.pack(side='right', fill='y')

        self._build_form(self.scroll_frame)

    def _build_form(self, parent):
        # ② 発注情報
        info = ttk.LabelFrame(parent, text='② 発注情報', padding=(8, 6))
        info.pack(fill='x', pady=(0, 6))

        self.field_vars: dict[str, tk.StringVar] = {}

        fields_left = [
            ('発注日',    'order_date',         datetime.date.today().isoformat()),
            ('通番号',    'serial_no',           ''),
            ('件名',      'subject',             ''),
            ('見積 No.',  'quote_no',            ''),
        ]
        _DELIVERY_LOCATIONS = [
            '福岡県北九州市小倉北区浅野二丁目14番3号あるあるcity2号館地下1階',
            '福岡県北九州市小倉北区浅野一丁目1番1号ビエラ小倉1F DISCOVERYcoworking-103',
        ]

        fields_right = [
            ('宛先（会社名）', 'maker_name',    ''),
            ('納期',          'delivery',      ''),
            ('支払条件',      'payment_terms', '月末締翌月末支払'),
        ]

        for side_frame, fields in [
            (ttk.Frame(info), fields_left),
            (ttk.Frame(info), fields_right),
        ]:
            side_frame.pack(side='left', fill='both', expand=True, padx=(0, 12))
            for i, (label, key, default) in enumerate(fields):
                ttk.Label(side_frame, text=label + ':', anchor='e', width=14).grid(
                    row=i, column=0, sticky='e', padx=(0, 4), pady=2)
                var = tk.StringVar(value=default)
                self.field_vars[key] = var
                w = 7 if key == 'serial_no' else 26
                ttk.Entry(side_frame, textvariable=var, width=w).grid(
                    row=i, column=1, sticky='w', pady=2)

        # 納品場所（ドロップダウン）
        right_frame = side_frame.master  # fields_right を入れた side_frame の親
        loc_row = len(fields_right)
        ttk.Label(side_frame, text='納品場所:', anchor='e', width=14).grid(
            row=loc_row, column=0, sticky='e', padx=(0, 4), pady=2)
        loc_var = tk.StringVar(value=_DELIVERY_LOCATIONS[0])
        self.field_vars['delivery_location'] = loc_var
        ttk.Combobox(
            side_frame, textvariable=loc_var,
            values=_DELIVERY_LOCATIONS, width=44, state='readonly'
        ).grid(row=loc_row, column=1, sticky='w', pady=2)

        # ③ 明細
        items_lf = ttk.LabelFrame(parent, text='③ 明細', padding=(8, 6))
        items_lf.pack(fill='both', expand=True, pady=(0, 6))

        self.items_frame = ttk.Frame(items_lf)
        self.items_frame.pack(fill='both', expand=True)
        self.items_frame.columnconfigure(1, weight=1)

        for j, (h, w, a) in enumerate(zip(
            ['No.', '摘要（品名）', '数量', '単価（税抜）', '金額', ''],
            [4,     38,            7,      11,             11,     4],
            ['center', 'w', 'center', 'center', 'e', 'center'],
        )):
            ttk.Label(self.items_frame, text=h, width=w, anchor=a,
                      font=('Yu Gothic UI', 9, 'bold')).grid(row=0, column=j, padx=2)

        ttk.Separator(self.items_frame, orient='horizontal').grid(
            row=1, column=0, columnspan=6, sticky='ew', pady=(2, 3))
        self._next_item_row = 2

        ttk.Button(items_lf, text='＋ 行を追加', command=self._add_item_row).pack(
            anchor='w', pady=(4, 0))

        # ④ 備考
        notes_lf = ttk.LabelFrame(parent, text='④ 備考', padding=(8, 6))
        notes_lf.pack(fill='x', pady=(0, 4))

        self.notes_text = tk.Text(notes_lf, height=4, font=('Yu Gothic UI', 9),
                                  relief='solid', borderwidth=1)
        self.notes_text.pack(fill='x')

    def _build_btn_bar(self, parent):
        bar = ttk.Frame(parent)
        bar.pack(fill='x')

        ttk.Button(bar, text='📊  Excel生成', command=self._gen_excel,
                   style='Action.TButton').pack(side='left', padx=(0, 8))
        ttk.Button(bar, text='📄  PDF生成',   command=self._gen_pdf,
                   style='Action.TButton').pack(side='left')

        self.total_var = tk.StringVar(value='小計: ¥-')
        ttk.Label(bar, textvariable=self.total_var,
                  font=('Yu Gothic UI', 10, 'bold'), foreground='#1a3a6b'
                  ).pack(side='right', padx=10)

    # ----------------------------------------------------------------
    # 初期値
    # ----------------------------------------------------------------

    def _populate_defaults(self):
        for _ in range(3):
            self._add_item_row()

    # ----------------------------------------------------------------
    # 明細行の追加・削除
    # ----------------------------------------------------------------

    def _add_item_row(self, description='', quantity='', unit_price=''):
        idx = self._next_item_row
        self._next_item_row += 1
        item = ItemRow(
            self.items_frame, idx,
            on_change=self._update_total,
            on_delete=lambda ri=idx: self._delete_item_row(ri),
            description=description, quantity=quantity, unit_price=unit_price,
        )
        self._item_rows.append(item)
        self._renumber_items()
        self._update_total()

    def _delete_item_row(self, row_idx: int):
        self._item_rows = [r for r in self._item_rows if r.row_idx != row_idx]
        for w in self.items_frame.grid_slaves():
            if w.grid_info().get('row') == row_idx:
                w.destroy()
        self._renumber_items()
        self._update_total()

    def _renumber_items(self):
        for i, r in enumerate(self._item_rows, 1):
            r.no_label.configure(text=str(i))

    def _update_total(self):
        total = sum(r.get_amount() for r in self._item_rows)
        tax   = int(total * 0.1)
        self.total_var.set(f'小計: ¥{total:,}　消費税: ¥{tax:,}　合計: ¥{total+tax:,}')

    # ----------------------------------------------------------------
    # PDF選択・プレビュー
    # ----------------------------------------------------------------

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title='メーカ見積書PDFを選択',
            filetypes=[('PDF files', '*.pdf'), ('All files', '*.*')]
        )
        if path:
            self.pdf_var.set(path)

    def _on_pdf_path_changed(self, *_):
        """パスが変わったら即プレビュー更新"""
        path = self.pdf_var.get().strip()
        if path and Path(path).is_file():
            self._preview.load(path)
        else:
            self._preview.clear()

    # ----------------------------------------------------------------
    # AI抽出
    # ----------------------------------------------------------------

    def _extract_async(self):
        pdf_path = self.pdf_var.get().strip()
        if not pdf_path:
            messagebox.showwarning('警告', 'PDFファイルを選択してください')
            return

        self._set_status('⏳ 抽出中... しばらくお待ちください', 'orange')
        self.update()

        def run():
            try:
                data = extract_from_pdf(pdf_path)
                self.after(0, lambda: self._on_extract_success(data))
            except Exception as e:
                self.after(0, lambda: self._on_extract_error(str(e)))

        threading.Thread(target=run, daemon=True).start()

    def _on_extract_success(self, data: dict):
        self._populate_form(data)
        self._set_status('✅ 抽出完了。内容を確認・修正してください', 'green')

    def _on_extract_error(self, msg: str):
        self._set_status(f'❌ エラー: {msg[:80]}', 'red')
        messagebox.showerror('抽出エラー', msg)

    def _set_status(self, msg: str, color: str = '#555'):
        self.status_var.set(msg)
        self.status_label.configure(foreground=color)

    def _populate_form(self, data: dict):
        for api_key in ('maker_name', 'quote_no', 'subject', 'delivery'):
            val = data.get(api_key, '')
            if val:
                self.field_vars[api_key].set(str(val))

        # 明細をリセット
        for r in list(self._item_rows):
            for w in self.items_frame.grid_slaves():
                if w.grid_info().get('row') == r.row_idx:
                    w.destroy()
        self._item_rows.clear()
        self._next_item_row = 2

        items = data.get('items', [])
        for item in (items if items else [{}] * 3):
            self._add_item_row(
                description=item.get('description', ''),
                quantity=item.get('quantity', ''),
                unit_price=item.get('unit_price', ''),
            )

        self.notes_text.delete('1.0', 'end')
        if data.get('notes'):
            self.notes_text.insert('1.0', data['notes'])

        self._update_total()

    # ----------------------------------------------------------------
    # フォームデータ取得
    # ----------------------------------------------------------------

    def _get_form_data(self) -> dict:
        try:
            order_date = datetime.date.fromisoformat(self.field_vars['order_date'].get())
        except ValueError:
            order_date = datetime.date.today()
        try:
            serial_no = int(self.field_vars['serial_no'].get())
        except ValueError:
            serial_no = 1

        return {
            'maker_name':        self.field_vars['maker_name'].get().strip(),
            'quote_no':          self.field_vars['quote_no'].get().strip(),
            'subject':           self.field_vars['subject'].get().strip(),
            'delivery':          self.field_vars['delivery'].get().strip(),
            'delivery_location': self.field_vars['delivery_location'].get().strip(),
            'payment_terms':     self.field_vars['payment_terms'].get().strip(),
            'order_date':        order_date,
            'serial_no':         serial_no,
            'items':             [r.to_dict() for r in self._item_rows if not r.is_empty()],
            'notes':             self.notes_text.get('1.0', 'end').strip(),
        }

    # ----------------------------------------------------------------
    # Excel・PDF生成
    # ----------------------------------------------------------------

    def _gen_excel(self):
        data = self._get_form_data()
        if not data['maker_name']:
            messagebox.showwarning('警告', '宛先（会社名）を入力してください')
            return
        path = make_output_path(data, '.xlsx')
        try:
            create_excel(data, path)
            self._set_status(f'✅ Excel生成: {path.name}', 'green')
            if messagebox.askyesno('完了', f'Excelを生成しました。\n\n{path}\n\n開きますか？'):
                os.startfile(path)
        except Exception as e:
            messagebox.showerror('エラー', str(e))

    def _gen_pdf(self):
        data = self._get_form_data()
        if not data['maker_name']:
            messagebox.showwarning('警告', '宛先（会社名）を入力してください')
            return
        excel_path = make_output_path(data, '.xlsx')
        pdf_path   = make_output_path(data, '.pdf')
        self._set_status('⏳ PDF生成中...', 'orange')
        self.update()
        try:
            create_excel(data, excel_path)
            excel_to_pdf(excel_path, pdf_path)
            self._set_status(f'✅ PDF生成: {pdf_path.name}', 'green')
            if messagebox.askyesno('完了', f'PDFを生成しました。\n\n{pdf_path}\n\n開きますか？'):
                os.startfile(pdf_path)
        except Exception as e:
            self._set_status(f'❌ エラー: {str(e)[:80]}', 'red')
            messagebox.showerror('エラー', str(e))


# ============================================================
# エントリポイント
# ============================================================

if __name__ == '__main__':
    app = App()
    app.mainloop()
